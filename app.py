import openpyxl as xl
from openpyxl.chart import BarChart, Reference

# turned it into a function to input filename instead of constantly changing source code
def process_workbook(filename):
    # create workbook by importing xlsx sheet
    wb = xl.load_workbook(filename)
    # we only have one sheet so set sheet to Sheet1
    sheet = wb['Sheet1']


    # for loop to start at 2 to skip header row
    for row in range(2, sheet.max_row + 1):
        # selecting cell for each loop
        cell = sheet.cell(row, 3)
        # creating new numbers off of new calculation
        corrected_cell = cell.value * 0.9
        # setting new cell to 2,4 3,4 4,4
        corrected_price_cell = sheet.cell(row, 4)
        # setting the corrected price to the new cell
        corrected_price_cell.value = corrected_cell

    # use the Reference class to grab values from desired rows and columns
    values = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=4, max_col=4)
    # create BarChart using BarChart class
    chart = BarChart()
    # add data values from Reference to the BarChart
    chart.add_data(values)
    # add the chart to the sheet at the cell 'E2'
    sheet.add_chart(chart, 'E2')
    # adds the new file to directory
    new = filename
    wb.save(new)
